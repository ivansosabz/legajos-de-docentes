from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

from .forms import DocenteForm, DocumentoForm
from .models import Docente, Documento, Notificacion
from datetime import date
from .utils import generar_notificaciones_vencidas
from .filters import DocenteFilter, DocumentoFilter

from django.contrib.auth.decorators import login_required, permission_required

import csv
import csv
import datetime
from openpyxl import Workbook

# Constantes
DOCENTES_POR_PAGINA = 5
DOCUMENTOS_POR_PAGINA = 6


# ------------------ DOCENTES ------------------
@login_required
def index(request):
    search_query = request.GET.get("search", "")

    # Base queryset
    docentes = Docente.objects.all().order_by("apellido", "nombre")

    # Aplicar búsqueda global si existe
    if search_query:
        docentes = docentes.filter(
            Q(nombre__icontains=search_query)
            | Q(apellido__icontains=search_query)
            | Q(ci__icontains=search_query)
            | Q(email__icontains=search_query)
        )

    # Aplicar filtros
    docente_filter = DocenteFilter(request.GET, queryset=docentes)
    docentes_filtrados = docente_filter.qs

    # Paginación
    paginator = Paginator(docentes_filtrados, 10)  # Ajusta el número por página
    page = request.GET.get("page")

    try:
        docentes_paginados = paginator.page(page)
    except PageNotAnInteger:
        docentes_paginados = paginator.page(1)
    except EmptyPage:
        docentes_paginados = paginator.page(paginator.num_pages)

    return render(
        request,
        "docentes/index.html",
        {
            "docentes": docentes_paginados,
            "filter": docente_filter,
            "search_query": search_query,
        },
    )


@login_required
def view(request, id):
    docente = get_object_or_404(Docente, id=id)
    return render(request, "docentes/details.html", {"docente": docente})


@login_required
@permission_required("docente.add_documento", raise_exception=True)
def create(request):
    form = DocenteForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Docente agregado correctamente.")
            return redirect("docente")
        messages.error(request, "Revisá los errores del formulario.")
    return render(request, "docentes/create.html", {"form": form})


@login_required
@permission_required("docente.change_documento", raise_exception=True)
def edit(request, id):
    docente = get_object_or_404(Docente, id=id)
    form = DocenteForm(request.POST or None, instance=docente)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Docente actualizado correctamente.")
        else:
            messages.error(request, "Revisá los errores del formulario.")
    return render(request, "docentes/edit.html", {"form": form, "id": id})


@login_required
@permission_required("docente.delete_documento", raise_exception=True)
@require_POST
def delete(request, id):
    docente = get_object_or_404(Docente, id=id)
    docente.delete()
    messages.success(request, "Docente eliminado correctamente.")
    return redirect("docente")


# ------------------ DOCUMENTOS ------------------
@login_required
def documento(request):
    # Obtener parámetros
    search_query = request.GET.get("search", "")

    # Base queryset
    documentos = Documento.objects.select_related("docente", "tipo_documento").order_by(
        "-fecha_emision"
    )

    # Aplicar búsqueda global si existe
    if search_query:
        documentos = documentos.filter(
            Q(nombre__icontains=search_query)
            | Q(docente__nombre__icontains=search_query)
            | Q(docente__apellido__icontains=search_query)
            | Q(docente__ci__icontains=search_query)
            | Q(tipo_documento__nombre__icontains=search_query)
        )

    # Aplicar filtros
    documento_filter = DocumentoFilter(request.GET, queryset=documentos)
    documentos_filtrados = documento_filter.qs

    # Paginación
    paginator = Paginator(documentos_filtrados, DOCUMENTOS_POR_PAGINA)
    page = request.GET.get("page")

    try:
        documentos_paginados = paginator.page(page)
    except PageNotAnInteger:
        documentos_paginados = paginator.page(1)
    except EmptyPage:
        documentos_paginados = paginator.page(paginator.num_pages)

    return render(
        request,
        "documentos/index.html",
        {
            "documentos": documentos_paginados,
            "filter": documento_filter,
            "search_query": search_query,
        },
    )


@login_required
@permission_required("docente.add_documento", raise_exception=True)
def create_document(request):
    form = DocumentoForm(request.POST or None, request.FILES or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Documento agregado correctamente.")
            return redirect("documento")
        messages.error(request, "Revisá los errores del formulario.")
    return render(request, "documentos/create.html", {"form": form})


@login_required
@permission_required("docente.change_documento", raise_exception=True)
def edit_document(request, id):
    documento = get_object_or_404(Documento, id=id)
    form = DocumentoForm(
        request.POST or None, request.FILES or None, instance=documento
    )

    if request.method == "POST":
        if form.is_valid():
            nuevo_doc = form.save()
            if (
                nuevo_doc.fecha_vencimiento
                and nuevo_doc.fecha_vencimiento > date.today()
            ):
                noti = Notificacion.objects.filter(
                    documento=nuevo_doc, estado="Pendiente"
                ).first()
                if noti:
                    noti.estado = "Renovado"
                    noti.fecha_resolucion = timezone.now()
                    noti.save()
            messages.success(request, "Documento actualizado correctamente.")
        else:
            messages.error(request, "Revisá los errores del formulario.")

    return render(request, "documentos/edit.html", {"form": form, "id": id})


@login_required
@permission_required("docente.delete_documento", raise_exception=True)
@require_POST
def delete_document(request, id):
    documento = get_object_or_404(Documento, id=id)
    documento.delete()
    messages.success(request, "Documento eliminado correctamente.")
    return redirect("documento")


@login_required
def notificacion(request):
    generar_notificaciones_vencidas()
    notificaciones = Notificacion.objects.filter(estado="Pendiente").order_by(
        "-fecha_envio"
    )
    return render(
        request, "notificaciones/index.html", {"notificaciones": notificaciones}
    )


# ------------------ REPORTES ------------------
@login_required
def reporte_documentos_form(request):
    """
    Muestra el formulario con select de docentes y rango de fechas.
    """
    docentes = Docente.objects.all().order_by(
        "apellido", "nombre"
    )  # ajusta si tu modelo tiene otros campos
    context = {
        "docentes": docentes,
    }
    return render(request, "docentes/reporte.html", context)


@login_required
def exportar_documentos(request):
    """
    Genera y devuelve el archivo (XLSX o CSV) con los documentos filtrados.
    Parámetros esperados via GET:
      - docente: id del docente o "all"
      - desde: fecha YYYY-MM-DD (opcional)
      - hasta: fecha YYYY-MM-DD (opcional)
      - formato: 'xlsx' o 'csv' (opcional, por defecto xlsx)
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    docente_id = request.GET.get("docente")
    fecha_desde = request.GET.get("desde")
    fecha_hasta = request.GET.get("hasta")
    formato = request.GET.get("formato", "xlsx").lower()

    qs = Documento.objects.all()

    if docente_id and docente_id != "all":
        qs = qs.filter(docente_id=docente_id)

    # detectar campos de fecha (robusto)
    fecha_carga_field = "fecha_emision" if hasattr(Documento, "fecha_emision") else None
    fecha_venc_field = (
        "fecha_vencimiento" if hasattr(Documento, "fecha_vencimiento") else None
    )

    if fecha_desde and fecha_carga_field:
        qs = qs.filter(**{f"{fecha_carga_field}__gte": fecha_desde})
    if fecha_hasta and fecha_carga_field:
        qs = qs.filter(**{f"{fecha_carga_field}__lte": fecha_hasta})

    qs = qs.select_related("docente").order_by(
        f"-{fecha_carga_field}" if fecha_carga_field else "-id"
    )

    # Construir filas con valores "raw" (fechas como objetos para XLSX)
    header = [
        "Fecha de Carga",
        "Docente",
        "Título / Descripción",
        "Fecha de Vencimiento",
    ]
    rows_raw = []
    for doc in qs:
        fecha_carga = getattr(doc, fecha_carga_field) if fecha_carga_field else None
        docente_obj = getattr(doc, "docente", None)
        nombre_docente = ""
        if docente_obj:
            nombre_docente = " ".join(
                filter(
                    None,
                    [
                        getattr(docente_obj, "apellido", ""),
                        getattr(docente_obj, "nombre", ""),
                    ],
                )
            ).strip()
        titulo = getattr(doc, "nombre", "")
        fecha_venc = getattr(doc, fecha_venc_field) if fecha_venc_field else None
        rows_raw.append([fecha_carga, nombre_docente, titulo, fecha_venc])

    today = datetime.date.today().isoformat()

    # CSV branch (no estilos) - formatear fechas como texto ISO
    if formato == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="reporte_documentos_{today}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(header)
        for row in rows_raw:

            def fmt(v):
                if v is None:
                    return ""
                if isinstance(v, (datetime.date, datetime.datetime)):
                    return v.strftime("%Y-%m-%d")
                return str(v)

            writer.writerow([fmt(c) for c in row])
        return response

    # XLSX branch (con estilo)
    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"

    # Escribir encabezado
    ws.append(header)

    # Escribir filas (fechas como objetos si las hay)
    for r in rows_raw:
        ws.append(r)

    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # azul suave
    thin_side = Side(border_style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Aplicar estilo al header
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Aplicar bordes y alineaciones a todas las celdas de datos y formatear fechas
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            # columna 1 y 4 son fechas (si vinieron como date/datetime)
            if idx in (1, 4):
                # si el valor es fecha (openpyxl mantiene objetos date), formatear
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell.number_format = "YYYY-MM-DD"
                    cell.alignment = center
                else:
                    cell.alignment = center
            else:
                cell.alignment = left

    # Ajustar ancho de columnas (puedes modificar valores)
    col_widths = {1: 15, 2: 30, 3: 60, 4: 15}
    for i in range(1, ws.max_column + 1):
        width = col_widths.get(i, 20)
        ws.column_dimensions[get_column_letter(i)].width = width

    # Filtro automático y congelar fila de encabezado
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws["A2"]

    # Responder
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_documentos_{today}.xlsx"'
    )
    wb.save(response)
    return response
